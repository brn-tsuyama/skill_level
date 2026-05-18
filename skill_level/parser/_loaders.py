"""
Sheet loader implementations (xls via xlrd, xlsx via openpyxl).

SheetLoader is a structural Protocol — any object with a matching load()
method qualifies, so new formats can be added without modifying existing code.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Protocol

import xlrd
from openpyxl import load_workbook

from skill_level.parser._models import _Sheet

if TYPE_CHECKING:
    from pathlib import Path


class SheetLoader(Protocol):
    """Load all worksheets from an Excel file into _Sheet objects."""

    def load(self, path: Path) -> list[_Sheet]: ...


class XlsLoader:
    def load(self, path: Path) -> list[_Sheet]:
        book = xlrd.open_workbook(str(path))
        return [
            _Sheet(name=sh.name, rows=[sh.row_values(r) for r in range(sh.nrows)])
            for sh in book.sheets()
        ]


class XlsxLoader:
    def load(self, path: Path) -> list[_Sheet]:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = [
            _Sheet(
                name=ws.title,
                rows=[list(row) for row in ws.iter_rows(values_only=True)],
            )
            for ws in wb.worksheets
        ]
        wb.close()
        return sheets


_LOADERS: dict[str, SheetLoader] = {
    ".xls": XlsLoader(),
    ".xlsx": XlsxLoader(),
}


def get_loader(path: Path) -> SheetLoader | None:
    """Return the appropriate loader for `path`, or None for unsupported extensions."""
    return _LOADERS.get(path.suffix.lower())
