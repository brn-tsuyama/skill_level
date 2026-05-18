"""Tests for skill_level.parser._loaders (XlsxLoader, get_loader)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from skill_level.parser._loaders import XlsxLoader, get_loader


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "06C022L11"
    ws.append(["能力ユニット名", "段取り"])
    ws.append(["概要", "型枠の段取り"])
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "様式２（施工技能）"
    ws2 = wb.create_sheet("06C022L11")
    ws2.append(["a", "b"])
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


class TestXlsxLoader:
    def test_loads_sheet_name(self, simple_xlsx: Path) -> None:
        loader = XlsxLoader()
        sheets = loader.load(simple_xlsx)
        assert len(sheets) == 1
        assert sheets[0].name == "06C022L11"

    def test_loads_cell_content(self, simple_xlsx: Path) -> None:
        loader = XlsxLoader()
        sheets = loader.load(simple_xlsx)
        assert sheets[0].rows[0][0] == "能力ユニット名"
        assert sheets[0].rows[0][1] == "段取り"

    def test_loads_multiple_sheets(self, multi_sheet_xlsx: Path) -> None:
        loader = XlsxLoader()
        sheets = loader.load(multi_sheet_xlsx)
        assert len(sheets) == 2
        assert sheets[0].name == "様式２（施工技能）"
        assert sheets[1].name == "06C022L11"

    def test_row_count(self, simple_xlsx: Path) -> None:
        loader = XlsxLoader()
        sheets = loader.load(simple_xlsx)
        assert sheets[0].nrows == 2


class TestGetLoader:
    def test_xlsx_extension(self, tmp_path: Path) -> None:
        path = tmp_path / "f.xlsx"
        loader = get_loader(path)
        assert loader is not None
        assert isinstance(loader, XlsxLoader)

    def test_xls_extension(self, tmp_path: Path) -> None:
        path = tmp_path / "f.xls"
        loader = get_loader(path)
        assert loader is not None

    def test_unknown_extension(self, tmp_path: Path) -> None:
        path = tmp_path / "f.csv"
        assert get_loader(path) is None

    def test_case_insensitive(self, tmp_path: Path) -> None:
        path = tmp_path / "f.XLSX"
        loader = get_loader(path)
        assert loader is not None
